import re, json
from openpyxl import Workbook, styles
from datetime import date

from constants import *


class Driver:

  def __init__(self) -> None:    
    print("initialized")
    self.wb = Workbook()
    # grab the active worksheet
    self.ws_client = self.wb.active
    self.current_row = ROW_START
    self.placement = json.load(open("./placement.json"))
    self.shorthand = json.load(open("./shorthand.json"))
    self.write_off_amount = 0
    self.write_off_members = 0

  def extract_items(self, text:str):
    lines = text.split("\n")
    first_line = lines.pop(0)
    first_line_items = re.findall(r"\w*\d*", first_line)
    first_line_items = [items for items in first_line_items if len(items)>0]
    # print(first_line_items)
    
    # TODO: improve row_prefix, support for west bengal --done
    state_name = ""
    state_idx = ""
    if(first_line_items[-1].isdigit()):
      state_name = ' '.join(first_line_items[:-1])
      state_idx = first_line_items[-1]
    else:
      state_name = ' '.join(first_line_items)
    
    row_prefix = self.shorthand[state_name.upper()] + ' ' + state_idx

    # row_prefix = self.shorthand["BIHAR"]
    print(row_prefix)
    for line in lines:
      groups = re.split("=|-|:|;|>|collection", maxsplit=1, string=line)
      # print(groups)
      bank_type = ""
      # lhs_split = groups[0].split() 
      lhs_split = re.findall("\w+", groups[0])
      for ele in lhs_split:
        element = ele.upper()
        if element in self.placement:
          bank_type = element + (" ADV" if "adv" in lhs_split else '')
          break
      
      if bank_type == "":
        if "bd" in lhs_split:
          items = re.findall("\w+", groups[1])
          self.write_off_amount += int(items[1] if len(items)>1 else 0)
          self.write_off_members += int(items[0])
        continue

      items = re.findall("\w+", groups[1])
      # print(items)
      item_placement = self.placement[bank_type]
      self.ws_client[f"{STATE_COLUMN}{self.current_row}"] = f'{row_prefix} {item_placement["row"]}'
      self.ws_client[f"{RECOVERY_COLUMN}{self.current_row}"] = items[2] if len(items)>2 else ''
      self.ws_client[f"{MEMBER_COLUMN}{self.current_row}"] = int(items[0])
      self.ws_client[f'{AMT_COLUMN[item_placement["col"]]}{self.current_row}'] = int(items[1])
      
      for col_idx in ['B', 'C', 'D', 'E', 'F']:
        self.ws_client[f"{col_idx}{self.current_row}"].alignment = styles.Alignment(horizontal='center')
        self.ws_client[f"{col_idx}{self.current_row}"].fill = styles.PatternFill(start_color=COLUMN_COLOR[col_idx], end_color=COLUMN_COLOR[col_idx], fill_type='solid')
        self.ws_client[f"{col_idx}{self.current_row}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))
      
      self.current_row += 1
      # print(f"type: {items[0]}, members: {items[1]} amount: {items[2]}, recovery: {items[3]}")

    pass

  def fill_total_cells(self):
    # SUB TOTAL for every BANKS
    for col_idx in ['C', 'D', 'E', 'F']:
      self.ws_client[f"{col_idx}{self.current_row}"] = f"=SUM({col_idx}{ROW_START}:{col_idx}{self.current_row-1})"
      self.ws_client[f"{col_idx}{self.current_row}"].alignment = styles.Alignment(horizontal='center')        
      self.ws_client[f"{col_idx}{self.current_row}"].fill = styles.PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid') 
      self.ws_client[f"{col_idx}{self.current_row}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))
    
    self.current_row += 1
    # TOTAL Collection header text
    self.ws_client.merge_cells(f"A{self.current_row-1}:B{self.current_row}")
    self.ws_client[f"A{self.current_row-1}"] = "Total Collection"
    self.ws_client[f"A{self.current_row-1}"].alignment = styles.Alignment(horizontal='center', vertical='center')        
    self.ws_client[f"A{self.current_row-1}"].fill = styles.PatternFill(start_color='FBE5D6', end_color='FBE5D6', fill_type='solid') 
    self.ws_client[f"A{self.current_row-1}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))  
    
    # TOTAL for all BANKS
    self.ws_client.merge_cells(f"C{self.current_row}:F{self.current_row}")
    self.ws_client[f"C{self.current_row}"] = f"=SUM(D{self.current_row-1}:F{self.current_row-1})"
    self.ws_client[f"C{self.current_row}"].alignment = styles.Alignment(horizontal='center')        
    self.ws_client[f"C{self.current_row}"].fill = styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') 
    self.ws_client[f"C{self.current_row}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))  

    pass

  def fill_header_cells(self):
    today = date.today()
    d1 = today.strftime("%d/%m/%Y")

    self.ws_client.merge_cells(f"A{ROW_START-2}:F{ROW_START-2}")
    self.ws_client[f"A{ROW_START-2}"] = f"STATE WISE RECOVERY FOR {d1}"      
    self.ws_client[f"A{ROW_START-2}"].alignment = styles.Alignment(horizontal='center')        
    self.ws_client[f"A{ROW_START-2}"].fill = styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') 
    self.ws_client[f"A{ROW_START-2}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))  

    self.ws_client[f"{STATE_COLUMN}{ROW_START-1}"] = "STATE"
    self.ws_client[f"{RECOVERY_COLUMN}{ROW_START-1}"] = "% RECOVERY"
    self.ws_client[f"{MEMBER_COLUMN}{ROW_START-1}"] = "MEMBERS"
    self.ws_client[f'{AMT_COLUMN["NBFC"]}{ROW_START-1}'] = "NBFC"
    self.ws_client[f'{AMT_COLUMN["SBI"]}{ROW_START-1}'] = "SBI"
    self.ws_client[f'{AMT_COLUMN["PRAYAAS"]}{ROW_START-1}'] = "PRAYAAS"

    for col_idx in ['A', 'B', 'C', 'D', 'E', 'F']:
      self.ws_client[f"{col_idx}{ROW_START-1}"].alignment = styles.Alignment(horizontal='center')        
      self.ws_client[f"{col_idx}{ROW_START-1}"].fill = styles.PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid') 
      self.ws_client[f"{col_idx}{ROW_START-1}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))
    
    pass

  def fill_write_off(self):
    self.current_row += 2
    # header
    self.ws_client.merge_cells(f"A{self.current_row}:C{self.current_row}")
    self.ws_client[f"A{self.current_row}"] = "TOTAL WRITE OFF OR OD COLLECTION"      
    self.ws_client[f"A{self.current_row}"].alignment = styles.Alignment(horizontal='center')        
    self.ws_client[f"A{self.current_row}"].fill = styles.PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid') 
    self.ws_client[f"A{self.current_row}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))  

    self.current_row += 1
    # w/o header
    self.ws_client.merge_cells(f"A{self.current_row}:A{self.current_row+1}")
    self.ws_client[f"A{self.current_row}"] = "W/O"      
    self.ws_client[f"A{self.current_row}"].alignment = styles.Alignment(horizontal='center', vertical='center')        
    self.ws_client[f"A{self.current_row}"].fill = styles.PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid') 
    self.ws_client[f"A{self.current_row}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))  

    # column sub-headers
    self.ws_client[f"B{self.current_row}"] = "MEMBERS"      
    self.ws_client[f"B{self.current_row}"].alignment = styles.Alignment(horizontal='center')        
    self.ws_client[f"B{self.current_row}"].fill = styles.PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid') 
    self.ws_client[f"B{self.current_row}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))  
    
    self.ws_client[f"C{self.current_row}"] = "AMOUNT"      
    self.ws_client[f"C{self.current_row}"].alignment = styles.Alignment(horizontal='center')        
    self.ws_client[f"C{self.current_row}"].fill = styles.PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid') 
    self.ws_client[f"C{self.current_row}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))  
    
    self.current_row += 1
    # write_off members, amount
    self.ws_client[f"B{self.current_row}"] = self.write_off_members      
    self.ws_client[f"B{self.current_row}"].alignment = styles.Alignment(horizontal='center')        
    self.ws_client[f"B{self.current_row}"].fill = styles.PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid') 
    self.ws_client[f"B{self.current_row}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))  
    
    self.ws_client[f"C{self.current_row}"] = self.write_off_amount      
    self.ws_client[f"C{self.current_row}"].alignment = styles.Alignment(horizontal='center')        
    self.ws_client[f"C{self.current_row}"].fill = styles.PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid') 
    self.ws_client[f"C{self.current_row}"].border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))  

    pass

  def extract(self, blocks):
    # set column width
    for col_idx in ['A', 'B', 'C']:
      self.ws_client.column_dimensions[col_idx].width = COLUMN_WIDTH

    self.fill_header_cells()
    
      # to maintain order
    print("\n\n@@@@@@@@@@@@@@@@@\nDRIVER", blocks)
    blocks.sort()
    # fill the table
    for element in blocks:
      if len(element)>0:
        self.extract_items(element)
    
    self.fill_total_cells()
    self.fill_write_off()
    print("END, saving...")
    
    self.wb.save("try.xlsx")

def entry_point():
  driver = Driver()
  blocks = []
  with open("./tmp.txt", "r") as f:
    text = f.read()
    terminal = "#"
    for t in text.split(terminal):
      if len(t)>0:
        blocks.append(t.strip().lower())

  driver.extract(blocks)
  pass

if __name__ == "__main__":
  entry_point()
